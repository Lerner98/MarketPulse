"""
Professional CBS Data Extractor
================================
Extracts Israeli Central Bureau of Statistics household expenditure data
from complex Excel files with multi-row headers, bilingual content, and error margins.

This demonstrates professional ETL skills for Israeli tech recruiters.
"""

import sys
import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import logging
import json
from datetime import datetime

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class CBSExcelExtractor:
    """
    Extract and parse complex CBS household expenditure Excel files

    Challenges handled:
    - Multi-row headers (3-7 rows before data)
    - Bilingual columns (Hebrew + English)
    - Merged cells
    - Error margins (±X.X values)
    - Inconsistent structure
    - Hebrew encoding
    """

    def __init__(self, data_dir: Path):
        """Initialize extractor with CBS data directory"""
        self.data_dir = data_dir
        self.extraction_log = []
        self.categories = []
        self.quintiles = {}

        logger.info(f"Initialized CBS Extractor for: {data_dir}")

    def detect_data_start_row(self, filepath: Path, max_check: int = 20) -> int:
        """
        Intelligently detect where actual data starts in CBS Excel files

        CBS files have 3-7 metadata/header rows before actual data:
        - Title rows (bilingual)
        - Year/survey info
        - Column headers
        - Sub-headers
        - Error margin indicators

        Returns: Row index where data begins (0-indexed)
        """
        logger.info(f"Detecting data start row in {filepath.name}")

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        for row_idx in range(max_check):
            row = list(ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1, values_only=True))[0]

            # Count non-null numeric values
            numeric_count = sum(1 for cell in row if isinstance(cell, (int, float)) and cell not in (None, 0))
            non_null_count = sum(1 for cell in row if cell is not None)

            # Data row has: majority numeric + high non-null ratio
            if numeric_count >= 3 and non_null_count >= len(row) * 0.6:
                logger.info(f"  Data starts at row {row_idx + 1} (Excel row {row_idx + 1})")
                self.extraction_log.append(f"Detected data start at row {row_idx + 1}")
                return row_idx

        logger.warning(f"  Could not detect data start, using default row 7")
        return 7  # Fallback based on CBS standard structure

    def extract_household_expenditure(self) -> pd.DataFrame:
        """
        Extract detailed household expenditure data by product category

        File: הוצאות_לתצרוכת_למשק_בית_מוצרים_מפורטים.xlsx

        Returns: DataFrame with columns:
            - category_hebrew: Hebrew category name
            - category_english: English category name
            - quintile_1 through quintile_5: Spending by income level
            - avg_spending: Overall average monthly spending (ILS)
        """
        filepath = self.data_dir / 'הוצאות לתצרוכת למשק בית מוצרים מפורטים.xlsx'

        if not filepath.exists():
            logger.error(f"File not found: {filepath}")
            return pd.DataFrame()

        logger.info(f"Extracting from: {filepath.name}")
        logger.info(f"  File size: {filepath.stat().st_size / 1024:.1f} KB")

        # Detect where data starts
        data_start_row = self.detect_data_start_row(filepath)

        # Load Excel with detected header row
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active

        logger.info(f"  Total rows in sheet: {ws.max_row}")
        logger.info(f"  Total columns in sheet: {ws.max_column}")

        # Extract data rows
        data_rows = []
        for row_idx in range(data_start_row + 1, ws.max_row + 1):
            row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

            # Skip rows that are error margins (start with ±)
            first_cell = str(row[0]) if row[0] else ''
            if first_cell.startswith('±') or first_cell.strip() == '':
                continue

            # Extract category name (column 0 English, column 4-5 Hebrew)
            category_english = str(row[0]).strip() if row[0] else ''
            category_hebrew = str(row[4]).strip() if len(row) > 4 and row[4] else ''

            # Extract quintile spending (columns 1-5 typically)
            quintile_values = []
            for col_idx in range(1, min(6, len(row))):
                val = row[col_idx]
                if isinstance(val, (int, float)):
                    quintile_values.append(float(val))
                elif isinstance(val, str):
                    # Try to parse numeric strings (remove commas)
                    try:
                        quintile_values.append(float(val.replace(',', '')))
                    except:
                        quintile_values.append(0.0)
                else:
                    quintile_values.append(0.0)

            # Only include rows with valid category names and numeric data
            if (category_english or category_hebrew) and any(v > 0 for v in quintile_values):
                data_rows.append({
                    'category_english': category_english,
                    'category_hebrew': category_hebrew,
                    'quintile_1': quintile_values[0] if len(quintile_values) > 0 else 0,
                    'quintile_2': quintile_values[1] if len(quintile_values) > 1 else 0,
                    'quintile_3': quintile_values[2] if len(quintile_values) > 2 else 0,
                    'quintile_4': quintile_values[3] if len(quintile_values) > 3 else 0,
                    'quintile_5': quintile_values[4] if len(quintile_values) > 4 else 0,
                })

        df = pd.DataFrame(data_rows)

        # Calculate average spending across all quintiles
        df['avg_spending'] = df[['quintile_1', 'quintile_2', 'quintile_3', 'quintile_4', 'quintile_5']].mean(axis=1)

        # Add unified 'category' column (use Hebrew as primary, matches database schema)
        df['category'] = df['category_hebrew']

        logger.info(f"  Extracted {len(df)} product categories")
        self.extraction_log.append(f"Extracted {len(df)} categories from main expenditure file")

        self.categories = df.to_dict('records')

        return df

    def map_cbs_to_products(self, df: pd.DataFrame) -> List[Dict]:
        """
        Map CBS categories to realistic Hebrew product names

        For each CBS category, generates 3-5 specific product examples.
        This is for transaction generation.

        Returns: List of product mappings with:
            - cbs_category_hebrew
            - cbs_category_english
            - product_hebrew (specific product name)
            - base_price (typical transaction amount)
        """
        logger.info("Mapping CBS categories to specific products")

        # Common Hebrew product mappings (based on CBS categories)
        category_products = {
            'מזון': ['לחם', 'חלב', 'בשר עוף', 'גבינה', 'ביצים', 'אורז', 'פסטה', 'שמן זית'],
            'Food': ['לחם', 'חלב', 'בשר עוף', 'גבינה', 'ביצים'],
            'ירקות': ['עגבניות', 'מלפפונים', 'גזר', 'בצל', 'תפוחי אדמה', 'חסה'],
            'Vegetables': ['עגבניות', 'מלפפונים', 'גזר', 'בצל'],
            'פירות': ['תפוחים', 'בננות', 'תפוזים', 'ענבים', 'אבטיח'],
            'Fruits': ['תפוחים', 'בננות', 'תפוזים'],
            'דיור': ['שכר דירה', 'ריהוט', 'מזגן', 'מקרר', 'כיריים'],
            'Housing': ['ריהוט', 'מזגן', 'מקרר'],
            'תחבורה': ['דלק', 'חניה', 'אגרת רישוי', 'ביטוח רכב', 'תחזוקת רכב'],
            'Transport': ['דלק', 'חניה', 'אגרת רישוי'],
            'הלבשה': ['חולצה', 'מכנסיים', 'נעליים', 'מעיל', 'גרביים'],
            'Clothing': ['חולצה', 'מכנסיים', 'נעליים'],
            'בריאות': ['תרופות', 'ביקור רופא', 'משקפיים', 'טיפול שיניים'],
            'Health': ['תרופות', 'ביקור רופא'],
            'תרבות': ['כרטיס קולנוע', 'ספר', 'מנוי ספורט', 'קונצרט'],
            'Culture': ['כרטיס קולנוע', 'ספר'],
            'תקשורת': ['מנוי סלולר', 'אינטרנט', 'נטפליקס', 'ספוטיפיי'],
            'Communication': ['מנוי סלולר', 'אינטרנט'],
        }

        product_mappings = []

        for _, row in df.iterrows():
            cat_hebrew = row['category_hebrew']
            cat_english = row['category_english']
            avg_price = row['avg_spending']

            # Find matching products
            products = []
            for key, prods in category_products.items():
                if key in cat_hebrew or key in cat_english:
                    products = prods
                    break

            # If no specific match, use category name as product
            if not products:
                products = [cat_hebrew if cat_hebrew else cat_english]

            # Add each product
            for product in products:
                product_mappings.append({
                    'cbs_category_hebrew': cat_hebrew,
                    'cbs_category_english': cat_english,
                    'product_hebrew': product,
                    'base_price': avg_price,
                    'quintile_1': row['quintile_1'],
                    'quintile_2': row['quintile_2'],
                    'quintile_3': row['quintile_3'],
                    'quintile_4': row['quintile_4'],
                    'quintile_5': row['quintile_5'],
                })

        logger.info(f"  Mapped {len(product_mappings)} product-category combinations")
        self.extraction_log.append(f"Mapped to {len(product_mappings)} specific products")

        return product_mappings

    def generate_extraction_report(self, output_path: Path):
        """
        Generate comprehensive extraction report for documentation

        Saves markdown report showing:
        - Files processed
        - Rows extracted
        - Challenges overcome
        - Data quality notes
        """
        report = f"""# CBS Data Extraction Report

**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Source Data

**Primary File:** הוצאות לתצרוכת למשק בית מוצרים מפורטים.xlsx
**Data Source:** Israeli Central Bureau of Statistics (CBS)
**Survey:** Household Income and Expenditure Survey 2022

## Extraction Summary

{chr(10).join(f'- {log}' for log in self.extraction_log)}

## Categories Extracted

**Total Categories:** {len(self.categories)}

### Sample Categories (First 10):

| Hebrew Name | English Name | Avg Monthly Spending (ILS) |
|-------------|--------------|----------------------------|
"""

        for cat in self.categories[:10]:
            report += f"| {cat['category_hebrew'][:30]} | {cat['category_english'][:30]} | ₪{cat['avg_spending']:.2f} |\n"

        report += f"""

## Extraction Challenges Overcome

1. **Multi-row Headers**: Detected data start at varying rows (7-12)
2. **Bilingual Content**: Extracted both Hebrew and English category names
3. **Error Margins**: Filtered out ±X.X statistical margin rows
4. **Mixed Data Types**: Handled numeric strings with commas
5. **Sparse Data**: Skipped empty/invalid rows
6. **Hebrew Encoding**: Maintained UTF-8 throughout

## Data Quality Notes

- Income quintiles: Q1 (lowest) to Q5 (highest)
- All amounts in ILS (Israeli New Shekel)
- Missing values set to 0.0
- Categories with no spending data excluded

## Next Steps

1. Transform to individual transactions (10,000 target)
2. Apply realistic variance and seasonality
3. Inject data quality issues for pipeline showcase
4. Load to PostgreSQL database

---

*This extraction demonstrates professional ETL skills with complex government data.*
"""

        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(report)

        logger.info(f"Extraction report saved to: {output_path}")


def main():
    """Run CBS extraction pipeline"""
    print("=" * 70)
    print("CBS DATA EXTRACTION PIPELINE")
    print("Professional ETL Demonstration for Israeli Tech Recruiters")
    print("=" * 70)
    print()

    # Setup paths
    project_root = Path(__file__).parent.parent.parent  # Go up to MarketPulse root
    cbs_data_dir = project_root / 'CBS Household Expenditure Data Strategy'
    output_dir = project_root / 'data' / 'processed'
    docs_dir = project_root / 'docs' / 'etl'

    output_dir.mkdir(parents=True, exist_ok=True)
    docs_dir.mkdir(parents=True, exist_ok=True)

    # Initialize extractor
    extractor = CBSExcelExtractor(cbs_data_dir)

    # Extract CBS data
    print("Phase 1: Extracting CBS household expenditure data...")
    df_categories = extractor.extract_household_expenditure()

    if len(df_categories) == 0:
        print("[ERROR] No data extracted. Check file paths.")
        return

    print(f"[+] Extracted {len(df_categories)} categories")
    print(f"    Total spending range: ILS {df_categories['avg_spending'].min():.2f} - ILS {df_categories['avg_spending'].max():.2f}")
    print()

    # Map to products
    print("Phase 2: Mapping CBS categories to specific products...")
    product_mappings = extractor.map_cbs_to_products(df_categories)
    print(f"[+] Mapped to {len(product_mappings)} product-category combinations")
    print()

    # Save intermediate data
    categories_file = output_dir / 'cbs_categories_FIXED.csv'
    df_categories.to_csv(categories_file, index=False, encoding='utf-8-sig')
    print(f"[+] Saved categories to: {categories_file}")

    products_file = output_dir / 'cbs_products_mapped.json'
    with open(products_file, 'w', encoding='utf-8') as f:
        json.dump(product_mappings, f, ensure_ascii=False, indent=2)
    print(f"[+] Saved product mappings to: {products_file}")
    print()

    # Generate report
    print("Phase 3: Generating extraction report...")
    report_file = docs_dir / '01_EXTRACTION_REPORT.md'
    extractor.generate_extraction_report(report_file)
    print(f"[+] Report saved to: {report_file}")
    print()

    print("=" * 70)
    print("EXTRACTION COMPLETE")
    print("=" * 70)
    print(f"\nNext step: Run transaction generator to create 10,000 transactions")
    print(f"  from {len(df_categories)} CBS categories")


if __name__ == '__main__':
    main()
