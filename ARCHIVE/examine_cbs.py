"""Examine CBS Excel file structure"""
import sys
import openpyxl
from pathlib import Path

# Force UTF-8 output
sys.stdout.reconfigure(encoding='utf-8')

cbs_file = Path('../CBS Household Expenditure Data Strategy/הוצאות לתצרוכת למשק בית מוצרים מפורטים.xlsx')

print(f'File: {cbs_file.name}')
print(f'Size: {cbs_file.stat().st_size / 1024:.1f} KB')
print('=' * 60)

wb = openpyxl.load_workbook(cbs_file, read_only=True, data_only=True)
print(f'Sheets: {len(wb.sheetnames)}')
ws = wb.active
print(f'Sheet name: {ws.title}')
print(f'Dimensions: {ws.max_row} rows x {ws.max_column} columns')
print('=' * 60)

print('\nFirst 15 rows:')
for i in range(1, min(16, ws.max_row + 1)):
    row = list(ws.iter_rows(min_row=i, max_row=i, values_only=True))[0]
    # Get first 6 columns
    cells = []
    for c in row[:6]:
        if c is None:
            cells.append('NULL')
        else:
            cells.append(str(c)[:25])
    print(f'Row {i:2d}: {cells}')

print('\n' + '=' * 60)
print(f'Total data rows available: {ws.max_row}')
